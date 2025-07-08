# from langchain_openai import ChatOpenAI
from langchain_google_genai import ChatGoogleGenerativeAI
from dataclasses import asdict, dataclass
from faker import Faker
import random
import os
from dotenv import load_dotenv
from pydantic import BaseModel
from pprint import pprint
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

load_dotenv()
faker = Faker()
NUM_ANIMALS = 5

@dataclass
class EvaluationResult:
    prompt: str
    llm_cat: str
    llm_dog: str
    ans_cat: str
    ans_dog: str
    accurate: bool

def llm_call() -> EvaluationResult:
    # llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash-lite-preview-06-17", api_key=os.getenv("GEMINI_API_KEY"))
    llm = ChatGoogleGenerativeAI(model="gemma-3-27b-it", api_key=os.getenv("GEMINI_API_KEY"))
    cats = [(faker.name(), i) for i in range(1, NUM_ANIMALS + 1)]
    dogs = [(faker.name(), i) for i in range(1, NUM_ANIMALS + 1)]

    cats_prompt = "\n".join([f"{name} is my {rank}th favorite cat" for name, rank in cats])
    dogs_prompt = "\n".join([f"{name} is my {rank}th favorite dog" for name, rank in dogs])

    random_cat_int = random.randint(1, NUM_ANIMALS)
    random_dog_int = random.randint(1, NUM_ANIMALS)

    fav_cat_answer = cats[random_cat_int - 1][0]
    least_fav_dog_answer = dogs[NUM_ANIMALS - random_dog_int][0]

    prompt = (
        f"You are a helpful agent who will answer which is my {random_cat_int}th favorite cat and {random_dog_int}th least favorite dog\n"
        f"{cats_prompt}\n"
        f"{dogs_prompt}\n"
        "\n"
        "Format the answer in json with cat and dog keys\n"
        "Example: {\"cat\": \"cat_name\", \"dog\": \"dog_name\"}\n"
        "Answer: "
    )

    class TestResult(BaseModel):
        cat: str
        dog: str

    # print(prompt)
    # llm_response = llm.with_structured_output(TestResult).invoke(prompt, config={"run_name": "purple-evals-1"})
    llm_response_str = llm.invoke(prompt, config={"run_name": "purple-evals-3"})
    # print(llm_response_str.content)
    try:
        llm_response = TestResult(**extract_json_from_fenced_block(llm_response_str.content)) # type: ignore
    except Exception as e:
        print(llm_response_str.content)
        print(f"Error: {e}")
        llm_response = TestResult(cat="", dog="")
    return EvaluationResult(
        prompt=prompt,
        llm_cat=llm_response.cat, # type: ignore
        llm_dog=llm_response.dog, # type: ignore
        ans_cat=fav_cat_answer,
        ans_dog=least_fav_dog_answer,
        accurate=llm_response.cat == fav_cat_answer and llm_response.dog == least_fav_dog_answer, # type: ignore
    )

def test_single_call():
    return llm_call()


def extract_json_from_fenced_block(input_str: str) -> dict:
    """
    Extracts JSON from a Markdown-style fenced code block like:
    \\```json
    {"cat": "Isabella Rocha", "dog": "Toni Hooper"}
    \\```
    """
    # Remove backticks and language label using regex
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', input_str, re.DOTALL)
    if match:
        json_str = match.group(1)
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            print("JSON decoding error:", e)
    return {}


def run_experiment(num_calls: int):
    results = []
    with ThreadPoolExecutor(max_workers=1) as executor:
        futures = {executor.submit(llm_call): None for _ in range(num_calls)}
        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(asdict(result))
                # pprint(result)
            except Exception as e:
                print(f"Error: {e}")
            # time.sleep(1)
    with open("results_gemma.json", "w") as f:
        json.dump(results, f)
        
    print(f"Accuracy: {sum([result['accurate'] for result in results]) / len(results)}")
    return results



def write_results_to_excel(json_path, excel_path):
    # Step 1: Load the JSON file
    with open(json_path, 'r') as f:
        data = json.load(f)

    # Step 2: Convert to DataFrame
    df = pd.DataFrame(data)

    # Step 3: Write to Excel
    df.to_excel(excel_path, index=False, sheet_name='Results')

    # Step 4: Format prompt column to wrap text
    wb = load_workbook(excel_path)
    ws = wb['Results']

    # Set word-wrap on the 'prompt' column (usually column 1)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Optionally, autofit column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 50)

    # Save changes
    wb.save(excel_path)

    

if __name__ == "__main__":
    # run_experiment(50)
    write_results_to_excel("results_gemma.json", "results_gemma.xlsx")